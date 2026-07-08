from datetime import date, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

from sqlalchemy import func, or_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from database.db import db
from models import Brand, Nota, NotaCustomer, NotaItem, NotaPayment, NotaProduct, SalesOrder
from services.number_generator import generate_nota_number
from utils.formatters import pretty_date


NOTA_STATUSES = (
    "Belum DP",
    "DP",
    "Lunas",
)

DEFAULT_NOTA_PRODUCTS = [
    ("FP", "Setelan full printing", 135000),
    ("LP1", "Tambahan lengan panjang", 15000),
    ("BE", "Bahan embosh", 5000),
    ("D30", "Diskon Rp30.000", -30000),
    ("DR", "Diskon regular", -10000),
    ("JA", "Diskon jersey anak", -10000),
    ("SL", "Diskon special line", -10000),
]

INVOICE_BRAND_EVPRO = {
    "key": "evpro",
    "display_name": "Evpro",
    "template_name": "Evpro",
}
INVOICE_BRAND_RDR = {
    "key": "evpro",
    "display_name": "RDR Apparel",
    "template_name": "Evpro",
}
INVOICE_BRAND_FF = {
    "key": "ff_apparel",
    "display_name": "FF Apparel",
    "template_name": "FF Apparel",
}


def list_brands():
    return Brand.query.filter_by(status="active").order_by(Brand.name).all()


def list_invoice_brand_options():
    defaults = ["Evpro", "RDR Apparel", "FF Apparel"]
    names = [brand.name for brand in Brand.query.order_by(Brand.name.asc()).all()]
    options = []
    for name in defaults + names:
        if name and name not in options:
            options.append(name)
    return options


def get_invoice_brand(brand_name):
    normalized = " ".join(str(brand_name or "").strip().casefold().split())
    if normalized == "ff apparel" or normalized == "ff":
        return INVOICE_BRAND_FF
    if normalized == "rdr apparel" or normalized == "rdr":
        return INVOICE_BRAND_RDR
    return INVOICE_BRAND_EVPRO


def get_invoice_brand_group(brand_name):
    normalized = " ".join(str(brand_name or "").strip().casefold().split())
    if normalized in ("ff apparel", "ff"):
        return "FF Apparel"
    if normalized in ("rdr apparel", "rdr"):
        return "RDR Apparel"
    return "EVPRO"


def invoice_brand_filter_options():
    return ["FF Apparel", "RDR Apparel", "EVPRO"]


def format_so_number_for_invoice(sales_order):
    if not sales_order:
        return "Manual"
    brand = sales_order.brand
    brand_name = brand.name if brand else ""
    brand_code = brand.code if brand else ""
    group = get_invoice_brand_group(brand_name or brand_code)
    so_number = sales_order.so_number or "-"
    if group in ("FF Apparel", "RDR Apparel"):
        return so_number
    parts = so_number.split("/", 1)
    if len(parts) != 2:
        return so_number
    prefix_source = brand_code or parts[0]
    abbreviation = _brand_abbreviation(prefix_source)
    return f"EV-{abbreviation}/{parts[1]}"


def format_invoice_so_code(sales_order):
    return format_so_number_for_invoice(sales_order)


def display_nota_number(nota):
    if nota and nota.sales_order and nota.sales_order.so_number:
        return f"Nota-{nota.sales_order.so_number}"
    return nota.nota_number if nota else "-"


def list_customers():
    return NotaCustomer.query.order_by(NotaCustomer.name).all()


def list_products():
    return NotaProduct.query.order_by(NotaProduct.code).all()


def upsert_product(form):
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    price = _parse_int(form.get("price"))
    if not code or not description:
        raise ValueError("Kode dan keterangan produk wajib diisi.")
    product = get_product_by_code(code)
    if not product:
        product = NotaProduct(code=code)
        db.session.add(product)
    product.description = description
    product.price = price
    db.session.commit()
    return product


def delete_product(product_id):
    product = NotaProduct.query.get_or_404(product_id)
    db.session.delete(product)
    db.session.commit()


def list_products_as_dicts():
    return [
        {"id": product.id, "code": product.code, "description": product.description, "price": product.price}
        for product in list_products()
    ]


def list_customers_as_dicts():
    return [
        {
            "id": customer.id,
            "brand_id": customer.brand_id,
            "name": customer.name,
            "team_name": customer.team_name,
            "phone": customer.phone,
            "address": customer.address,
        }
        for customer in list_customers()
    ]


def list_notas(search=None):
    search = search or {}
    query = Nota.query.join(NotaCustomer).join(Brand).outerjoin(SalesOrder, Nota.so_id == SalesOrder.id)
    q = str(search.get("q") or "").strip()
    status = str(search.get("status") or "").strip()
    brand_id = str(search.get("brand_id") or "").strip()
    brand_group = str(search.get("brand_group") or "").strip()

    if q:
        normalized_q = q.removeprefix("Nota-")
        needle = f"%{q}%"
        normalized_needle = f"%{normalized_q}%"
        query = query.filter(
            or_(
                Nota.nota_number.ilike(needle),
                NotaCustomer.name.ilike(needle),
                Nota.team_name.ilike(needle),
                NotaCustomer.phone.ilike(needle),
                SalesOrder.so_number.ilike(needle),
                SalesOrder.so_number.ilike(normalized_needle),
            )
        )
    if brand_id.isdigit():
        query = query.filter(Nota.brand_id == int(brand_id))
    notas = query.order_by(Nota.order_date.desc(), Nota.id.desc()).all()
    if brand_group in invoice_brand_filter_options():
        notas = [nota for nota in notas if get_invoice_brand_group(nota.brand.name if nota.brand else "") == brand_group]
    if status:
        notas = [nota for nota in notas if calculate_invoice_status(nota) == status]
    return notas


def nota_rows(search=None):
    return [_nota_row(nota) for nota in list_notas(search)]


def report_nota_rows(brand=None):
    return [_nota_row(nota) for nota in _filtered_notas(brand)]


def get_nota(nota_id):
    return Nota.query.get_or_404(nota_id)


def delete_nota(nota):
    db.session.delete(nota)
    db.session.commit()


def get_nota_by_so_id(so_id):
    if not str(so_id or "").isdigit():
        return None
    return Nota.query.filter_by(so_id=int(so_id)).first()


def get_product_by_code(code):
    return NotaProduct.query.filter(func.upper(NotaProduct.code) == str(code or "").strip().upper()).first()


def totals(nota):
    return {"total": nota.total, "paid": nota.paid, "remaining": nota.remaining}


def calculate_invoice_status(nota):
    total = int(nota.total or 0)
    paid = int(nota.paid or 0)
    remaining = max(total - paid, 0)
    if paid <= 0:
        return "Belum DP"
    if paid >= total or remaining <= 0:
        return "Lunas"
    return "DP"


def invoice_status_badge_class(status):
    return {
        "Belum DP": "status-belum-dp",
        "DP": "status-dp",
        "Lunas": "status-lunas",
    }.get(status, "status-belum-dp")


def dashboard_stats(brand=None):
    notas = _filtered_notas(brand)
    revenue = sum(nota.total for nota in notas)
    income = sum(nota.paid for nota in notas)
    return _row(
        revenue=revenue,
        income=income,
        receivable=revenue - income,
        invoice_count=len(notas),
        belum_dp_count=sum(1 for nota in notas if calculate_invoice_status(nota) == "Belum DP"),
        dp_count=sum(1 for nota in notas if calculate_invoice_status(nota) == "DP"),
        lunas_count=sum(1 for nota in notas if calculate_invoice_status(nota) == "Lunas"),
        desain_count=0,
        produksi_count=0,
        selesai_count=0,
        diambil_count=0,
    )


def monthly_revenue(brand=None):
    buckets = {}
    for nota in _filtered_notas(brand):
        key = nota.order_date.strftime("%Y-%m") if nota.order_date else "-"
        buckets[key] = buckets.get(key, 0) + nota.total
    return [_row(month=month, total=total) for month, total in sorted(buckets.items())]


def yearly_revenue(brand=None):
    buckets = {}
    for nota in _filtered_notas(brand):
        key = nota.order_date.strftime("%Y") if nota.order_date else "-"
        buckets[key] = buckets.get(key, 0) + nota.total
    return [_row(year=year, total=total) for year, total in sorted(buckets.items(), reverse=True)]


def top_customers(brand=None):
    rows = {}
    for nota in _filtered_notas(brand):
        customer = nota.customer
        key = customer.id
        row = rows.setdefault(
            key,
            {
                "name": customer.name,
                "team_name": nota.team_name,
                "phone": customer.phone,
                "brand": nota.brand.name if nota.brand else "-",
                "invoice_count": 0,
                "order_count": 0,
                "total": 0,
            },
        )
        row["invoice_count"] += 1
        row["order_count"] += 1
        row["total"] += nota.total
    sorted_rows = sorted(rows.values(), key=lambda row: (-row["invoice_count"], -row["total"], row["name"]))
    return [_row(**row) for row in sorted_rows[:10]]


def receivables(brand=None, status=None):
    rows = []
    for nota in _filtered_notas(brand):
        if status and calculate_invoice_status(nota) != status:
            continue
        row = _nota_row(nota)
        if row.remaining > 0:
            rows.append(row)
    return rows


def income_payments(brand=None):
    query = NotaPayment.query.join(Nota).join(NotaCustomer)
    if brand:
        query = query.join(Brand, Nota.brand_id == Brand.id).filter(Brand.name == brand)
    payments = query.order_by(NotaPayment.payment_date.desc(), NotaPayment.id.desc()).all()
    return [
        _row(
            payment_date=payment.payment_date,
            amount=payment.amount,
            description=payment.description,
            invoice_number=display_nota_number(payment.nota),
            brand=payment.nota.brand.name if payment.nota.brand else "-",
            customer_name=payment.nota.customer.name,
            team_name=payment.nota.team_name,
        )
        for payment in payments
    ]


def income_summary(brand=None):
    today = date.today()
    week_start = today - timedelta(days=6)
    month_key = today.strftime("%Y-%m")
    year_key = today.strftime("%Y")
    payments = income_payments(brand)
    return _row(
        today=sum(payment.amount for payment in payments if payment.payment_date == today),
        weekly=sum(payment.amount for payment in payments if payment.payment_date and payment.payment_date >= week_start),
        monthly=sum(payment.amount for payment in payments if payment.payment_date and payment.payment_date.strftime("%Y-%m") == month_key),
        yearly=sum(payment.amount for payment in payments if payment.payment_date and payment.payment_date.strftime("%Y") == year_key),
    )


def workbook_response(sheet_name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="C5162E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 3, 45)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def invoice_export_rows(rows):
    return [
        [
            row.invoice_number,
            pretty_date(row.order_date),
            row.brand,
            row.customer_name,
            row.team_name,
            row.total,
            row.paid,
            row.remaining,
            row.status,
        ]
        for row in rows
    ]


def item_rows_for_form(nota=None):
    if not nota:
        return []
    return [
        {
            "product_code": item.product_code,
            "quantity": item.quantity,
            "description": item.description,
            "price": item.price,
            "subtotal": item.subtotal,
        }
        for item in sorted(nota.items, key=lambda item: item.sort_order)
    ]


def posted_item_rows(form):
    codes = form.getlist("product_code[]")
    quantities = form.getlist("quantity[]")
    rows = []
    for index, code in enumerate(codes):
        rows.append(
            {
                "product_code": code,
                "quantity": quantities[index] if index < len(quantities) else "",
            }
        )
    return rows


def form_data_from_sales_order(sales_order):
    customer_access = sales_order.customer_access
    order_date = sales_order.created_at.date() if sales_order.created_at else date.today()
    return {
        "so_id": sales_order.id,
        "brand_id": sales_order.brand_id,
        "order_date": order_date.isoformat(),
        "status": "Belum DP",
        "customer_name": customer_access.customer_name if customer_access else sales_order.team_name,
        "team_name": sales_order.team_name,
        "phone": customer_access.customer_phone if customer_access else "",
        "address": "",
        "notes": f"Dibuat dari Sales Order {sales_order.so_number}",
    }


def item_rows_from_sales_order(sales_order):
    total_size = sales_order.total_size
    if total_size <= 0:
        return []
    if not get_product_by_code("FP"):
        return []
    return [{"product_code": "FP", "quantity": total_size}]


def billing_status_for_sales_order(sales_order):
    nota = get_nota_by_so_id(sales_order.id)
    if not nota:
        return "Belum Ada Nota"
    status = calculate_invoice_status(nota)
    if status == "Lunas":
        return "Lunas"
    if status == "DP":
        return "DP"
    return "Nota Dibuat"


def parse_item_rows(form):
    codes = form.getlist("product_code[]")
    quantities = form.getlist("quantity[]")
    items = []
    errors = []
    for index, code in enumerate(codes, start=1):
        code = str(code or "").strip().upper()
        quantity_raw = str(quantities[index - 1] if index - 1 < len(quantities) else "").strip()
        if not code and not quantity_raw:
            continue
        if not code:
            errors.append(f"Baris {index}: kode produk harus diisi.")
            continue
        if not quantity_raw.isdigit() or int(quantity_raw) <= 0:
            errors.append(f"Baris {index}: qty harus angka lebih dari 0.")
            continue
        product = get_product_by_code(code)
        if product is None:
            errors.append(f"Baris {index}: kode produk {code} belum ada di database.")
            continue
        items.append(
            {
                "product_id": product.id,
                "code": product.code,
                "description": product.description,
                "price": product.price,
                "quantity": int(quantity_raw),
            }
        )
    if not items:
        errors.append("Minimal satu item produk harus diisi.")
    return items, errors


def validate_nota_form(form):
    errors = []
    if not _get_brand(form.get("brand_id")):
        errors.append("Brand wajib dipilih.")
    if not str(form.get("order_date") or "").strip():
        errors.append("Tanggal order wajib diisi.")
    elif not _parse_date(form.get("order_date")):
        errors.append("Tanggal order tidak valid.")
    if not str(form.get("customer_name") or "").strip():
        errors.append("Nama customer wajib diisi.")
    if not str(form.get("team_name") or "").strip():
        errors.append("Nama tim wajib diisi.")
    _, item_errors = parse_item_rows(form)
    errors.extend(item_errors)
    return errors


def create_nota(form, user=None):
    items, errors = parse_item_rows(form)
    if errors:
        raise ValueError("\n".join(errors))
    order_date = _parse_date(form.get("order_date"))
    customer = _save_customer(form)
    so_id = _parse_so_id(form.get("so_id"))
    existing_nota = get_nota_by_so_id(so_id) if so_id else None
    if existing_nota:
        return existing_nota
    nota = Nota(
        nota_number=generate_nota_number(order_date),
        brand_id=int(form.get("brand_id")),
        order_date=order_date,
        customer=customer,
        team_name=str(form.get("team_name") or "").strip(),
        status="Belum DP",
        notes=str(form.get("notes") or "").strip() or None,
        so_id=so_id,
        created_by_id=user.id if user else None,
    )
    _replace_items(nota, items)
    db.session.add(nota)
    db.session.commit()
    return nota


def update_nota(nota, form):
    items, errors = parse_item_rows(form)
    if errors:
        raise ValueError("\n".join(errors))
    customer = _save_customer(form)
    nota.brand_id = int(form.get("brand_id"))
    nota.order_date = _parse_date(form.get("order_date"))
    nota.customer = customer
    nota.team_name = str(form.get("team_name") or "").strip()
    nota.notes = str(form.get("notes") or "").strip() or None
    so_id = _parse_so_id(form.get("so_id"))
    if so_id and so_id != nota.so_id:
        existing_nota = get_nota_by_so_id(so_id)
        if existing_nota:
            raise ValueError("Sales Order ini sudah memiliki Nota.")
    nota.so_id = so_id
    _replace_items(nota, items)
    db.session.commit()
    return nota


def add_payment(nota, form):
    payment_date = _parse_date(form.get("payment_date"))
    amount = _parse_int(form.get("amount"))
    if not payment_date:
        raise ValueError("Tanggal pembayaran wajib diisi.")
    if amount <= 0:
        raise ValueError("Nominal pembayaran harus lebih dari 0.")
    db.session.add(
        NotaPayment(
            nota=nota,
            payment_date=payment_date,
            amount=amount,
            description=str(form.get("description") or "").strip() or None,
        )
    )
    db.session.commit()


def seed_default_nota_products():
    existing = {product.code.upper(): product for product in NotaProduct.query.all()}
    for code, description, price in DEFAULT_NOTA_PRODUCTS:
        product = existing.get(code.upper())
        if product:
            product.description = description
            product.price = price
        else:
            db.session.add(NotaProduct(code=code, description=description, price=price))


def _filtered_notas(brand=None):
    query = Nota.query.join(Brand)
    if brand:
        query = query.filter(Brand.name == brand)
    return query.order_by(Nota.order_date.desc(), Nota.id.desc()).all()


def _nota_row(nota):
    status = calculate_invoice_status(nota)
    invoice_number = display_nota_number(nota)
    return _row(
        id=nota.id,
        invoice_number=invoice_number,
        nota_number=invoice_number,
        brand=nota.brand.name if nota.brand else "-",
        order_date=nota.order_date,
        status=status,
        status_class=invoice_status_badge_class(status),
        customer_name=nota.customer.name,
        team_name=nota.team_name,
        phone=nota.customer.phone,
        total=nota.total,
        paid=nota.paid,
        remaining=nota.remaining,
    )


def _row(**kwargs):
    return SimpleNamespace(**kwargs)


def _save_customer(form):
    brand_id = int(form.get("brand_id"))
    customer_id = form.get("customer_id")
    customer = NotaCustomer.query.get(int(customer_id)) if str(customer_id or "").isdigit() else None
    if not customer:
        customer = NotaCustomer()
        db.session.add(customer)
    customer.brand_id = brand_id
    customer.name = str(form.get("customer_name") or "").strip()
    customer.team_name = str(form.get("team_name") or "").strip()
    customer.phone = str(form.get("phone") or "").strip() or None
    customer.address = str(form.get("address") or "").strip() or None
    return customer


def _replace_items(nota, items):
    nota.items.clear()
    for index, item in enumerate(items, start=1):
        nota.items.append(
            NotaItem(
                product_id=item["product_id"],
                product_code=item["code"],
                description=item["description"],
                price=int(item["price"]),
                quantity=int(item["quantity"]),
                subtotal=int(item["price"]) * int(item["quantity"]),
                sort_order=index,
            )
        )


def _get_brand(brand_id):
    if not str(brand_id or "").isdigit():
        return None
    return Brand.query.get(int(brand_id))


def _parse_date(value):
    if isinstance(value, date):
        return value
    value = str(value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _parse_so_id(value):
    if not str(value or "").isdigit():
        return None
    sales_order = SalesOrder.query.filter_by(id=int(value), is_deleted=False).first()
    return sales_order.id if sales_order else None


def _brand_abbreviation(value):
    value = "".join(char for char in str(value or "").strip().upper() if char.isalnum())
    if not value:
        return "ERP"
    if value == "EVPRO":
        return "EVP"
    return value[:3]
